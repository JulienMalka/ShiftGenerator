import openpyxl
from pylatex import Document, Section, Subsection, Command
from pylatex.utils import italic, NoEscape
from operator import itemgetter
import re


texts = {}
texts["Chalet"] = "Cette tâche comprend la prise en main du chalet, son aménagement avant que les participants arrivent et l'état des lieux, ainsi que son rengement et nettoyage intégral à la fin du weekend."
texts["ChaletR"] = "Ils sont en charge du contact avec les propriétaires et ont la chagre d'attribuer des tâches aux staffs qu'ils ont sous la main si celles-ci ne sont pas déjà définies; EN PLUS du travail de staff classique."

texts["Bus"] = "Reception des participants Avenue Piccard à 16h, les faire monter dans le bus ainsi que leurs affaires, le plus rapidement possible sans oublier personne. Départ avant 16h30 de Lausanne impératif. Vous êtes aussi en charge de l'ambiance pendant le voyage. Vous donnez le ton pour la suite du weekend !"
texts["BusR"] = "Ils auront avec eux les listes des participants. Leur rôle principal est de contrôler la monté dans le bus et faire l'appel affin de s'assurer que tout le monde embarque; EN PLUS du travail de staff classique."

texts["PrepApero"] = "Sont chargés d'installer l'apéro, préparer les boissons et la nourriture ainsi que de ranger l'apéro."
texts["PrepAperoR"] = "Simple rôle de supervision. En plus du travail de staff classique."


texts["PrepRepas"] = "Sont chargés de préparer la cuisine et d'organiser l'installation du couvert (peuvent demander de l'aide à des staffs qui n'ont pas de tâche attribuée), ainsi que de ranger et nettoyer les tables."

texts["PrepRepasR"] = "Ont la lourde tâche de préparer de bons repas tout en respectant des timing assez serrés. Ont les pleins pouvoirs sur leurs staffs. En plus du travail de staff classique."

texts["PrepSoiree"] = "Sont chargés de construire les jeux, d'aménager la salle, les lights et la sono et de préparer les boissons."
texts["PrepSoireeR"] = "Simple rôle de supervision. En plus du travail de staff classique."

texts["Vaisselle"] = "Tout le monde voit de quoi je parle je pense."
texts["VaisselleR"] = "Doivent s'assurer que la vaisselle est faite dans les bons timing; EN PLUS du travail de staff classique."

texts["Bar"] = "S'assurer que personne de trop cuit se mette mal. Pas la peine de donner les boissons aux participants, ils peuvent se servir. Aussi en charge de reload les boissons."
texts["BarR"] = "Simple rôle de supervision. En plus du travail de staff classique."


texts["NettoyageSoiree"] = "Rôle ingrat mais primordial, permet de réduir très fortement le travail à faire dimanche au moment de rendre le chalet. Ici plus qu'ailleurs le but est d'être rapide et efficace."
texts["NettoyageSoireeR"] = "Ont le rôle d'impulser un rythme de travail soutenu, pour assurer le plus de sommeil aux staffs."

texts["PrepPetitDejeuner"] = "Rôle difficile de part son aspect très matinal. J'ai fait en sorte que ceux qui doivent se lever tôt n'aient pas de shifts après 01h du matin, mais il ne tient qu'à vous de vous coucher ou faire la fête jusqu'à 4h. Cependant rien de vous sera pardonné au moment de taffer. En effet, la prep du petit dej comprend : installer et ranger les tables et la nourriture du p'tit dej, ainsi que préparer les 120 sandwichs pour le midi. C'est sans doute le rôle le plus important du weeekend. Voilà donc pourquoi vous ne devez rien laisser au hasard !!"

texts["PrepPetitDejeunerR"] = "Les timing sont très serrés pour cette tâche, les responsables ont un réel rôle de métronome. Vous avez là aussi les pleins pouvoir sur vos stafs, il est primordial que l'organisation soit ultra-efficace."


book = openpyxl.load_workbook('shifts.xlsx')
sheet = book.active

VENDREDI = 11
SAMEDI = 35
DIMANCHE = 53

tasks_v = []
tasks_s = []
tasks_d = []


NB_COLUMN = 12

searched = str(120)


def numbertoname(number):
    return sheet.cell(row=number-100+2,column=15).value

def stringchange(test):
    lol = test.replace(" ","")
    return lol.replace("é","e")

def taskwith(numbers):
    numbers = re.findall('\d+', numbers)
    string = ""
    for number in numbers:
        string = string  + numbertoname(int(number)) + ', '
    return string[:-2]+"."

def populate(start, end):
    hours=0
    tasks = []
    for i in range(1, NB_COLUMN):
        for j in range(start, end):
            cell = str(sheet.cell(row=j, column=i).value)

            if searched in cell or "TOUT LE MONDE" in cell:
                if "R"+ searched in cell or "R "+searched in cell:
                    hours = hours+1
                    status = 3
                elif "TOUT LE MONDE" in cell:
                    status = 1
                else:
                    hours = hours+1
                    status=2
                tasks.append((str(sheet.cell(row=1, column=i).value), str(sheet.cell(row=j,column=2).value),status,sheet.cell(row=j,column=i).value))
    h = 0
    for i in tasks:
        print(i[0])
    print("------------------------------")

    task = tasks[0][0]
    start = int(tasks[0][1][:2])
    starti = 0;
    status = tasks[0][2]
    end = start +1
    tasks_complete = []

    while h < len(tasks):
        #print(tasks[h][0])
        if task != tasks[h][0] or h+1==len(tasks) or status !=tasks[h][2] or (h>0 and int(tasks[h][1][:2])!= int(tasks[h-1][1][:2])+1):
            if h>0 and h+1 == len(tasks) and task != tasks[h][0]:

                tasks_complete.append((task, start,int(tasks[h-1][1][:2])+1, tasks[h-1][2], tasks[starti][3]))
                tasks_complete.append((tasks[h][0], int(tasks[h][1][:2]),int(tasks[h][1][:2])+1, tasks[h][2], tasks[h][3]))
            elif int(tasks[h][1][:2])!= int(tasks[h-1][1][:2])+1:
                tasks_complete.append((tasks[h-1][0], start,int(tasks[h-1][1][:2])+1, tasks[h-1][2], tasks[starti][3]))

            elif h+1==len(tasks):
                tasks_complete.append((task, start,int(tasks[h][1][:2])+1, tasks[h][2], tasks[starti][3]))
            elif task!=tasks[h][0]:
                tasks_complete.append((task, start, end, tasks[h-1][2], tasks[starti][3]))

            status = tasks[h][2]
            start = int(tasks[h][1][:2])
            starti = h
            end = start+1
            task = tasks[h][0]


        else:
            end = int(tasks[h][1][:2])+1
        h = h+1

    return (sorted(tasks_complete, key=itemgetter(1)), hours)



ven = (populate(1,VENDREDI+1))
#print(ven)
sam = populate(VENDREDI+1, SAMEDI+1)
dim = populate(SAMEDI+1, DIMANCHE+1)

total = ven[1]+sam[1] + dim[1]


ven = ven[0]
sam = sam[0]
dim = dim[0]

#FILE GENERATION

doc = Document('basic')

doc.preamble.append(Command('title', 'Récapitulatif Personnel Week-End Ski 2018 : '+str(numbertoname(int(searched)))))
doc.preamble.append(Command('date', NoEscape('')))
doc.preamble.append(NoEscape(r'\usepackage{xcolor}'))
doc.append(NoEscape(r'\maketitle'))
doc.append(NoEscape(r" \textbf{Nombre d'heures : }" + str(total)))



with doc.create(Section('Vendredi', False)):
    for i in ven:
         #print (i[4])
         if(i[3]==2 or i[3]==3):
             with doc.create(Subsection(str(i[0])+ " : " +str(i[1])+ "h - "+str(i[2])+"h, avec "+taskwith(i[4]), False)):
                  doc.append("Description : "+texts[stringchange(i[0])])
                  doc.append(NoEscape(r'\newline'))
                  if(i[3]==3):
                      doc.append(NoEscape(r'\textcolor{red}{Attention, vous êtes également responsable pour cette tâche : }'+texts[stringchange(i[0])+"R"]))


with doc.create(Section('Samedi',False)):
    for i in sam:
         if(i[3]==2 or i[3]==3):
             with doc.create(Subsection(str(i[0])+ " : " +str(i[1])+ "h - "+str(i[2])+"h, avec "+taskwith(i[4]), False)):
                  doc.append("Description : "+texts[stringchange(i[0])])
                  doc.append(NoEscape(r'\newline'))
                  if(i[3]==3):
                      doc.append(NoEscape(r'\textcolor{red}{Attention, vous êtes également responsable pour cette tâche : }'+texts[stringchange(i[0])+"R"]))

with doc.create(Section('Dimanche', False)):
    for i in dim:
         if(i[3]==2 or i[3]==3):
             with doc.create(Subsection(str(i[0])+ " : " +str(i[1])+ "h - "+str(i[2])+"h, avec "+taskwith(i[4]), False)):
                  doc.append("Description : "+texts[stringchange(i[0])])
                  doc.append(NoEscape(r'\newline'))
                  if(i[3]==3):
                      doc.append(NoEscape(r'\textcolor{red}{Attention, vous êtes également responsable pour cette tâche : }'+texts[stringchange(i[0])+"R"]))


with doc.create(Section('Autres tâches :',False)):
    doc.append("Comme tout le monde, vous devez également vous rendre disponible pour aider dans ces tâches :")
    doc.append(NoEscape(r'\begin{itemize}'))
    for i in ven:
        if(i[3]==1):
            doc.append(NoEscape(r'\item Vendredi de '+ str(i[1])+ 'h à '+str(i[2])+'h : '+str(i[0])))

    for i in sam:
        if(i[3]==1):
            doc.append(NoEscape(r'\item Samedi de '+ str(i[1])+ 'h à '+str(i[2])+'h : '+str(i[0])))

    for i in dim:
        #print(i[3])
        if(i[3]==1):
            doc.append(NoEscape(r'\item Dimanche de '+ str(i[1])+ 'h à '+str(i[2])+'h : '+str(i[0])))

    doc.append(NoEscape(r'\end{itemize}'))





doc.generate_pdf('shift_'+numbertoname(int(searched)), clean_tex=False)
